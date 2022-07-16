import csv

from openpyxl import Workbook
from openpyxl.styles import Alignment

import es_data20220402
from es_data import miss_from_type_server, miss_from_type_client


def parse_data(buckets1):
    res_lst = []
    for bucket in buckets1:
        row_lst = []
        # print(bucket["key"], bucket["doc_count"])
        flow_name = bucket["key"]
        flow_req = bucket["doc_count"]
        sever_ip_lst = []
        for ip in bucket["IP_agg"]["buckets"]:
            sever_ip = ip["key"]
            sever_ip_lst.append(sever_ip)
        sever_ip_str = "\n".join(sever_ip_lst)
        row_lst.append(flow_name)
        row_lst.append(sever_ip_str)
        row_lst.append(" ")
        row_lst.append(flow_req)
        print(row_lst)
        res_lst.append(row_lst)
    with open("data.csv", "a", newline='', encoding='utf-8-sig') as csv_file:
        writer = csv.writer(csv_file)

        for row in res_lst:
            writer.writerow(row)


def parse_data1(buckets1):
    res_lst = []
    for bucket in buckets1:
        row_lst = []
        # print(bucket["key"], bucket["doc_count"])
        flow_name = bucket["key"]
        flow_req = bucket["doc_count"]
        sever_ip_lst = []
        # for ip in bucket["my_aggs1"]["buckets"]:
        #     sever_ip = ip["key"]
        #     sever_ip_lst.append(sever_ip)
        # sever_ip_str = "\n".join(sever_ip_lst)
        row_lst.append(flow_name)
        # row_lst.append(sever_ip_str)
        # row_lst.append(" ")
        row_lst.append(flow_req)
        print(row_lst)
        res_lst.append(row_lst)
    with open("data.csv", "a", newline='', encoding='utf-8-sig') as csv_file:
        writer = csv.writer(csv_file)

        for row in res_lst:
            writer.writerow(row)


'''
{
          "key" : "Flow2KBMS",
          "doc_count" : 7788887,
          "Command" : {
            "doc_count_error_upper_bound" : 0,
            "sum_other_doc_count" : 0,
            "buckets" : [
              {
                "key" : "9.4.13.40",
                "doc_count" : 1902920
              },
'''


def parse_data2(buckets):
    res_lst = []
    res_lst2 = []
    for bucket in buckets:
        row_lst = []
        flow_name = bucket["key"]
        flow_req = bucket["doc_count"]

        sever_ip_lst = []
        sever_ip_lst2 = []
        for ip in bucket["Command"]["buckets"]:
            sever_ip = ip["key"]
            sever_ip_becall = ip["key"] + "|" + str(ip["doc_count"])
            sever_ip_lst.append(sever_ip_becall)
            sever_ip_lst2.append(sever_ip)

        sever_ip_str = "\n".join(sever_ip_lst)
        sever_ip_str2 = "|".join(sever_ip_lst2)
        row_lst.append(flow_name)
        row_lst.append(sever_ip_str)
        row_lst.append(flow_req)

        # 索引0为flowname ,后面为flow的serverip
        res_lst2.append(flow_name + "|" + sever_ip_str2 + "\n")
        res_lst.append(row_lst)
    with open("xml_becall_data2.csv", "a", newline='', encoding='utf-8-sig') as csv_file:
        writer = csv.writer(csv_file)
        for row in res_lst:
            writer.writerow(row)
    with open("xml_and_server_ip.txt", "w", encoding="utf8") as f:
        for line in res_lst2:
            f.writelines(line)


# 为了使得格式更加规范,现用openpyxl重写
def pares_data_to_xlsx(buckets):
    book = Workbook()
    sheet = book.active
    row = 1
    col = 1
    merge_i = 1

    line_lis = []
    for bucket in buckets:
        flow_name = bucket["key"]
        flow_req = bucket["doc_count"]
        sheet.cell(row, col, flow_name)
        sheet.cell(row, col + 1, flow_req)
        # 测试服务器排除 先记录总量位置
        flow_req_r = row
        flow_req_c = col + 1

        ips = []
        for ip in bucket["Command"]["buckets"]:

            sever_ip = ip["key"]
            sever_ip_count = ip["doc_count"]

            # 测试服务器排除
            if sever_ip in ["9.138.64.30", "9.138.64.37"]:
                flow_req -= sever_ip_count
                sheet.cell(flow_req_r, flow_req_c, flow_req)
                continue

            sheet.cell(row, col + 2, sever_ip)
            sheet.cell(row, col + 3, sever_ip_count)
            row += 1
            ips.append(sever_ip)
        line = flow_name + "|" + "|".join(ips) + "\n"
        line_lis.append(line)
        # 间隔大于1才需要合并
        if row - 1 - merge_i >= 1:
            # flow_name
            sheet.merge_cells(start_row=merge_i, end_row=row - 1, start_column=1, end_column=1)
            print(f"merge_start:{merge_i},merge_end:{row - 1}")
            # flow_count
            sheet.merge_cells(start_row=merge_i, end_row=row - 1, start_column=2, end_column=2)
        merge_i = row
        print("无需合并")
    book.save("so_becall_data.xlsx")

    with open("so_and_server_ip.txt", "w", encoding="utf8") as f:
        for line in line_lis:
            f.writelines(line)


if __name__ == '__main__':
    # parse_data(buckets1)
    # parse_data(buckets2)
    # parse_data1(from_type_web)
    # parse_data(miss_from_type_server)
    # parse_data(miss_from_type_client)
    # parse_data2(es_data20220402.all_xml)
    pares_data_to_xlsx(es_data20220402.all_so)