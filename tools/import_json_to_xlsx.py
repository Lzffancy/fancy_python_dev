from openpyxl import Workbook

raw_data = [{"key": "CmdRedisWX", "doc_count": 80701466}, {"key": "CmdL5", "doc_count": 80222087}, {"key": "CmdCurlSo", "doc_count": 17923453}, {"key": "CmdRedisNewComm", "doc_count": 16825819}, {"key": "CmdSSDbWx", "doc_count": 14045209},
            {"key": "CmdGenOpenidInfo", "doc_count": 13609252}, {"key": "CmdWxInfo", "doc_count": 11615243}, {"key": "CmdGetGameInfo", "doc_count": 11524198}, {"key": "CmdMysqlSo", "doc_count": 9434254},
            {"key": "CmdGameProgramDesOpenid", "doc_count": 9356399}, {"key": "CmdMSDKCheck", "doc_count": 7875015}, {"key": "CmdWxInfoNew", "doc_count": 7666144}, {"key": "CmdMiniProgramDesOpenid", "doc_count": 6269536},
            {"key": "CmdOidb", "doc_count": 6219492}, {"key": "CmdGetWxInfo", "doc_count": 5366724}, {"key": "CmdTof", "doc_count": 5044029}, {"key": "CmdTcpSend", "doc_count": 4388192}, {"key": "CmdCurlSoKG", "doc_count": 3420351},
            {"key": "CmdXYUPS", "doc_count": 3231394}, {"key": "CmdRedisComm", "doc_count": 3117101}, {"key": "CmdSSDbCon", "doc_count": 3107866}, {"key": "CmdYPProgramDesOpenid", "doc_count": 2832343},
            {"key": "CmdKFWxPayInfo", "doc_count": 2780928}, {"key": "CmdGetUserInfoIvr", "doc_count": 2607452}, {"key": "CmdCreatTask", "doc_count": 1870897}, {"key": "CmdGetPostOptInfo", "doc_count": 1863228},
            {"key": "CmdKGPunish", "doc_count": 1725111}, {"key": "CmdQQMiniProgramDesOpenid", "doc_count": 1626790}, {"key": "CmdAes", "doc_count": 1602340}, {"key": "CmdKGPenaltyNew", "doc_count": 1554294},
            {"key": "CmdCftInfo", "doc_count": 1468464}, {"key": "CmdReportProgramDesOpenid", "doc_count": 1300572}, {"key": "CmdCftHKWallet", "doc_count": 1032644}, {"key": "CmdCaf", "doc_count": 1016836},
            {"key": "CmdSapmg", "doc_count": 987350}, {"key": "CmdChannel", "doc_count": 927156}, {"key": "CmdTowerPkg", "doc_count": 615390}, {"key": "CmdGetGameOdpInfo", "doc_count": 603685},
            {"key": "CmdWxEnterTempsession", "doc_count": 533092}, {"key": "CmdTowerBan", "doc_count": 522819}, {"key": "CmdAppealQQ", "doc_count": 514014}, {"key": "CmdTcpString", "doc_count": 502029},
            {"key": "CmdSendWxMsg", "doc_count": 479542}, {"key": "CmdGetGameLoopInfo", "doc_count": 383147}, {"key": "CmdPayAntiFraudComm", "doc_count": 325496}, {"key": "CmdQQHead", "doc_count": 324483},
            {"key": "CmdQQGroupAppeal", "doc_count": 310300}, {"key": "CmdMobileRegistration", "doc_count": 309229}, {"key": "CmdGetGameKFPunishRecord", "doc_count": 224103}, {"key": "CmdAPPCallH5", "doc_count": 209767}]
book = Workbook()
sheet = book.active
row = 1
col = 1
for data in raw_data:
    command = data["key"]
    count = data["doc_count"]

    sheet.cell(row, col, command)
    sheet.cell(row, col + 1, count)
    row += 1
    book.save("so_top50_data.xlsx")
